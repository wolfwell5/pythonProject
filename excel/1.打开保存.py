from openpyxl import Workbook, load_workbook

# wb = Workbook()
# ws = wb.active
# print(ws.title)
#
# wb.save('E:\SoftWareDevelop\Repos\pythonProject1st\\test.xlsx')


wb = load_workbook("E:\Develop\Repos\pythonProject\excel\\files\\test.xlsx")
ws = wb.active
print(ws.title)

ws.title = "661"
wb.save(r"E:\Develop\Repos\pythonProject\excel\files\test.xlsx")
