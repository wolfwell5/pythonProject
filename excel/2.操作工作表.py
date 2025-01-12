from openpyxl import Workbook


wb = Workbook()
ws = wb.active
print(ws.title)

ws2 = wb.create_sheet("no.2",1)
ws3 = wb.create_sheet("no.3",2)

print(wb.sheetnames)

wb.move_sheet(ws3, -1)

del wb["no.2"]

print(wb.sheetnames)

