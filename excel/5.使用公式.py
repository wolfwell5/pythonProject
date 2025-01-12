from openpyxl import Workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import FORMULAE

# print(len(FORMULAE))
# print("SUM" in FORMULAE)


wb = Workbook()
ws = wb.active

ws.append(['价格1', '价格2', '平均值', '合计'])
ws.append([1, 3])
ws.append([2, 4])
ws.append([3, 5])
ws.append([4, 6])

ws['c2'] = '=sum(a2:b2)'
ws['d2'] = '=average(a2:b2)'

# ws['c3'] = Translator(formula='=sum(a2:b2)', origin='c2').translate_formula('c3')
# ws['c4'] = Translator(formula='=sum(a3:b3)', origin='c3').translate_formula('c4')
#
# ws['d3'] = Translator(formula='=average(a2:b2)', origin='d2').translate_formula('d3')
# ws['d4'] = Translator(formula='=average(a2:b2)', origin='d2').translate_formula('d4')

for cell in ws['c4:c5']:
    cell[0].value = Translator(formula='=average(a2:b2)', origin='c2').translate_formula(cell[0].coordinate)

wb.save('5.使用公式.xlsx')
