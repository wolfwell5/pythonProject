from openpyxl import Workbook

wb = Workbook()
ws = wb.active

x = 1
for i in range(1, 6):
    for j in range(1, 6):
        ws.cell(i, j, x)
        x = x + 1

# wb.save("3.单元格操作.xlsx")

# for i in range(1, 6):
#     for j in range(1, 6):
# print(ws.cell(i, j))

# print("in ws[a1:b2]", ws["a1:b2"])
# print()
#
# for row in ws["a1:b2"]:
#     print(row)
#     for cell in row:
#         print(cell.value)


print(ws["a:c"])
print()

print(ws["1:3"])
print()

print(ws["a1:c4"])
print()

print(ws["a"])
